from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


AMOUNT_COL = (340, 540)
TOP_AMOUNT_COL = (470, 570)
TOTAL_COL = (1360, 1585)


@dataclass
class Row:
    y: float
    words: list[dict]

    @property
    def text(self) -> str:
        return " ".join(w["text"] for w in self.words)


def clean_text(s: str) -> str:
    return re.sub(r"\s+", "", s)


def parse_amount_token(token: str) -> int | None:
    raw = token.strip()
    if not raw:
        return None
    if raw.startswith(",") or raw.startswith("."):
        return None
    negative = raw.startswith(("-", "•", "·", "ㆍ", "“", "”", "‘", "'"))
    raw = raw.replace("O", "0").replace("o", "0").replace("이", "0")
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return None
    value = int(digits)
    return -value if negative else value


def row_amount(row: Row, x_min: int, x_max: int) -> int | None:
    parts: list[str] = []
    for word in row.words:
        x = word["box"]["x"]
        if x_min <= x <= x_max:
            txt = word["text"]
            if re.search(r"[\d,.\-•·ㆍ]", txt):
                parts.append(txt)
    if not parts:
        return None

    parsed = [parse_amount_token(p) for p in parts]
    parsed = [p for p in parsed if p is not None]
    if len(parsed) == 1:
        return parsed[0]

    joined = "".join(parts)
    joined_value = parse_amount_token(joined)
    if joined_value is not None:
        return joined_value
    return parsed[0] if parsed else None


def rows_in_band(rows: list[Row], y_min: float, y_max: float) -> list[Row]:
    return [row for row in rows if y_min <= row.y <= y_max]


def amount_in_band(rows: list[Row], y_min: float, y_max: float, x_min: int, x_max: int) -> int | None:
    for row in rows_in_band(rows, y_min, y_max):
        value = row_amount(row, x_min, x_max)
        if value is not None:
            return value
    return None


def total_in_bottom_rows(rows: list[Row], x_min: int, x_max: int) -> int | None:
    candidates: list[tuple[float, int]] = []
    for row in rows:
        if row.y < 1180:
            continue
        compact = clean_text(row.text)
        if ("합" in compact or "함" in compact) and "계" in compact:
            value = row_amount(row, x_min, x_max)
            if value is not None:
                candidates.append((row.y, value))
    if candidates:
        return sorted(candidates)[-1][1]
    return amount_in_band(rows, 1260, 1325, x_min, x_max)


def detail_negative_refund(detail_rows: list[Row]) -> int | None:
    for row in rows_in_band(detail_rows, 295, 350):
        value = row_amount(row, 280, 470)
        if value is not None and value <= 0:
            return value
    return None


def structured_non_sol_values(detail_page: dict | None, detail_ratio: float, full_rows: list[Row]) -> dict[str, int | None]:
    if detail_page is None:
        return {}
    detail_rows = grouped_rows(detail_page, detail_ratio)
    values = {
        "electric": amount_in_band(detail_rows, 40, 90, 280, 470),
        "tv": amount_in_band(detail_rows, 80, 125, 280, 470),
        "parking": amount_in_band(detail_rows, 170, 230, 280, 470),
        "water": amount_in_band(detail_rows, 260, 320, 280, 470),
        "refund": detail_negative_refund(detail_rows),
        "total": total_in_bottom_rows(detail_rows, 280, 470),
    }

    # Cropped-table OCR sometimes drops the minus sign on refund rows; the full-page
    # OCR keeps it more reliably because it sees the whole green delta column too.
    full_refund = find_value(full_rows, match_water_refund, lower=True, col=AMOUNT_COL)
    if full_refund is not None and full_refund < 0:
        values["refund"] = full_refund
    return values


def structured_sol_values(rows: list[Row]) -> dict[str, int | None]:
    return {
        "electric": amount_in_band(rows, 250, 310, 430, 580),
        "water": amount_in_band(rows, 295, 340, 430, 580),
        "tv": amount_in_band(rows, 335, 380, 430, 580),
        "parking": None,
        "refund": None,
        "total": amount_in_band(rows, 640, 710, TOTAL_COL[0], TOTAL_COL[1]),
    }


def first_non_none(*values):
    for value in values:
        if value is not None:
            return value
    return None


def normalized_word(word: dict, coord_ratio: float, x_offset: float = 0, y_offset: float = 0) -> dict:
    if coord_ratio == 1 and x_offset == 0 and y_offset == 0:
        return word
    box = word["box"]
    return {
        **word,
        "box": {
            "x": box["x"] / coord_ratio + x_offset,
            "y": box["y"] / coord_ratio + y_offset,
            "width": box["width"] / coord_ratio,
            "height": box["height"] / coord_ratio,
        },
    }


def grouped_rows(page: dict, coord_ratio: float = 1, x_offset: float = 0, y_offset: float = 0) -> list[Row]:
    words: list[dict] = []
    for line in page["lines"]:
        for word in line["words"]:
            words.append(normalized_word(word, coord_ratio, x_offset, y_offset))
    words.sort(key=lambda w: (w["box"]["y"], w["box"]["x"]))

    rows: list[list[dict]] = []
    for word in words:
        y = word["box"]["y"]
        if rows and abs(sum(w["box"]["y"] for w in rows[-1]) / len(rows[-1]) - y) <= 18:
            rows[-1].append(word)
        else:
            rows.append([word])
    return [Row(sum(w["box"]["y"] for w in row) / len(row), sorted(row, key=lambda w: w["box"]["x"])) for row in rows]


def apt_name(text: str) -> str:
    compact = clean_text(text)
    if "건산주공" in compact:
        return "건산주공아파트"
    if "코아루해피트리" in compact:
        return "코아루해피트리"
    if "솔뫼" in compact or "솔외" in compact:
        return "솔뫼타운아파트"
    match = re.search(r"([가-힣A-Za-z0-9]+아파트)", text)
    return match.group(1) if match else ""


def dong_ho(text: str) -> tuple[str, str]:
    compact = clean_text(text)
    match = re.search(r"(\d{3})동(\d{1,4})호", compact)
    if match:
        return match.group(1), match.group(2)
    dong = re.search(r"(\d{3})동", compact)
    ho = re.search(r"(\d{1,4})호", compact)
    dong_value = dong.group(1) if dong else ""
    ho_value = ho.group(1) if ho else ""
    pay_no = re.search(r"납부번호[:：,]?(\d{4})-\d+-(\d{4})", compact)
    if pay_no:
        if not ho_value:
            ho_value = pay_no.group(1).lstrip("0") or "0"
        if not dong_value:
            dong_value = pay_no.group(2).lstrip("0") or "0"
    return dong_value, ho_value


def match_electric(row: Row) -> bool:
    t = clean_text(row.text).lower()
    return ("전기" in t and ("kw" in t or "k\\" in t or "k넓" in t)) or "전기료(" in t


def match_tv(row: Row) -> bool:
    t = clean_text(row.text)
    return "수신료" in t or ("수신" in t and "료" in t)


def match_water(row: Row) -> bool:
    t = clean_text(row.text)
    if "공동수도" in t:
        return False
    return ("수도료(" in t) or ("수" in t and "도" in t and re.search(r"m|㎥|耐", row.text.lower()) is not None)


def match_water_refund(row: Row) -> bool:
    t = clean_text(row.text)
    return "수도" in t and ("잉여환급" in t or "임여환급" in t or "환급" in t)


def match_parking(row: Row) -> bool:
    return "주차" in clean_text(row.text)


def find_value(rows: list[Row], predicate, *, lower: bool, col=AMOUNT_COL) -> int | None:
    selected: Iterable[Row] = rows
    if lower:
        selected = [r for r in rows if 950 <= r.y <= 2300]
    else:
        selected = [r for r in rows if r.y <= 900]
    for row in selected:
        if predicate(row):
            value = row_amount(row, *col)
            if value is not None:
                return value
    return None


def find_non_sol_total(rows: list[Row]) -> int | None:
    for row in rows:
        if row.y < 1900:
            continue
        if ("합" in row.text or "함" in row.text) and "계" in row.text:
            value = row_amount(row, *AMOUNT_COL)
            if value is not None:
                return value
    return None


def find_sol_total(rows: list[Row]) -> int | None:
    for row in rows:
        t = clean_text(row.text)
        if row.y <= 900 and ("납기내" in t or "남기내" in t or ("기내" in t and "금액" in t)):
            value = row_amount(row, *TOTAL_COL)
            if value is not None:
                return value
    return None


def fallback_total_from_text(text: str) -> int | None:
    due_match = re.search(r"금액은\s*([\d,]{5,})\s*원", text)
    if due_match:
        return int(due_match.group(1).replace(",", ""))
    amounts = [int(x.replace(",", "")) for x in re.findall(r"\b\d{2,3},\d{3}\b", text)]
    bill_like = [amount for amount in amounts if 50_000 <= amount <= 300_000]
    return max(bill_like or amounts) if amounts else None


def page_by_number(data: dict) -> dict[int, dict]:
    return {page["page"]: page for page in data.get("pages", [])}


def extract_page(page: dict, coord_ratio: float = 1, detail_page: dict | None = None, detail_ratio: float = 1) -> dict:
    text = page["text"]
    rows = grouped_rows(page, coord_ratio)
    apt = apt_name(text)
    is_sol = apt.startswith("솔뫼")
    dong, ho = dong_ho(text)

    lower = not is_sol
    col = AMOUNT_COL if lower else TOP_AMOUNT_COL
    structured = structured_sol_values(rows) if is_sol else structured_non_sol_values(detail_page, detail_ratio, rows)

    fallback_electric = find_value(rows, match_electric, lower=lower, col=col)
    fallback_tv = find_value(rows, match_tv, lower=lower, col=col)
    fallback_water = find_value(rows, match_water, lower=lower, col=col)
    fallback_refund = find_value(rows, match_water_refund, lower=lower, col=col)
    fallback_parking = find_value(rows, match_parking, lower=lower, col=col)

    electric = first_non_none(structured.get("electric"), fallback_electric)
    tv = first_non_none(structured.get("tv"), fallback_tv)
    water = first_non_none(structured.get("water"), fallback_water)
    refund = first_non_none(structured.get("refund"), fallback_refund)
    parking = first_non_none(structured.get("parking"), fallback_parking)

    if not is_sol:
        if electric is None:
            electric = find_value(rows, match_electric, lower=not lower, col=TOP_AMOUNT_COL if lower else AMOUNT_COL)
        if tv is None:
            tv = find_value(rows, match_tv, lower=not lower, col=TOP_AMOUNT_COL if lower else AMOUNT_COL)
        if water is None:
            water = find_value(rows, match_water, lower=not lower, col=TOP_AMOUNT_COL if lower else AMOUNT_COL)
        if parking is None:
            parking = find_value(rows, match_parking, lower=not lower, col=TOP_AMOUNT_COL if lower else AMOUNT_COL)

    compact = clean_text(text).lower()
    if ("수도0m3" in compact or "수도0m" in compact or "수도0" in compact):
        water = 0

    water_net = None
    if water is not None:
        water_net = water + (refund or 0)

    total = first_non_none(structured.get("total"), find_sol_total(rows) if is_sol else find_non_sol_total(rows))
    if total is None:
        total = fallback_total_from_text(text)

    return {
        "페이지": page["page"],
        "아파트명": apt,
        "동": dong,
        "호수": ho,
        "TV수신료": tv,
        "수도료": water_net,
        "주차비": parking,
        "전기료": electric,
        "납기내총액": total,
        "수도기본금액": water,
        "수도환급금": refund,
    }


def fmt(value):
    return "" if value is None else value


def main() -> None:
    if len(sys.argv) not in (3, 4):
        raise SystemExit("usage: build_management_fee_xlsx.py <ocr_json> <out_xlsx> [detail_ocr_json]")
    data = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8-sig"))
    out_xlsx = Path(sys.argv[2])
    detail_data = json.loads(Path(sys.argv[3]).read_text(encoding="utf-8-sig")) if len(sys.argv) == 4 else None
    detail_pages = page_by_number(detail_data) if detail_data else {}

    render_scale = data.get("render", {}).get("scale", 2)
    coord_ratio = render_scale / 2 if render_scale else 1
    detail_ratio = render_scale / 2 if render_scale else 1
    extracted = [extract_page(page, coord_ratio, detail_pages.get(page["page"]), detail_ratio) for page in data["pages"]]

    wb = Workbook()
    ws = wb.active
    ws.title = "추출결과"
    headers = ["아파트명", "동", "호수", "TV수신료", "수도료", "주차비", "전기료", "납기내총액", "페이지"]
    ws.append(headers)
    for item in extracted:
        ws.append([fmt(item[h]) for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center" if cell.column in (2, 3, 9) else "right")
            if 4 <= cell.column <= 8 and isinstance(cell.value, int):
                cell.number_format = '#,##0'
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    audit = wb.create_sheet("검토용")
    audit.append(["페이지", "수도기본금액", "수도환급금", "OCR원문"])
    for item, page in zip(extracted, data["pages"]):
        audit.append([item["페이지"], fmt(item["수도기본금액"]), fmt(item["수도환급금"]), page["text"]])
    for cell in audit[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    audit.column_dimensions["D"].width = 100
    for row in audit.iter_rows(min_row=2, max_col=4):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")

    for sheet in wb.worksheets:
        for idx, width in enumerate([18, 8, 8, 12, 12, 12, 12, 14, 8, 14, 14], start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    print(out_xlsx)


if __name__ == "__main__":
    main()
