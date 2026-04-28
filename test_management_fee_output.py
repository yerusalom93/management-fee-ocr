from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    headers = [cell.value for cell in ws[1]]
    return [
        {headers[idx]: value for idx, value in enumerate(row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    ]


def main() -> int:
    if len(sys.argv) != 3:
        print("usage: test_management_fee_output.py <xlsx> <expected_json>")
        return 2

    actual = load_rows(Path(sys.argv[1]))
    expected = json.loads(Path(sys.argv[2]).read_text(encoding="utf-8-sig"))
    mismatches = []
    for expected_row, actual_row in zip(expected, actual, strict=False):
        page = expected_row.get("페이지")
        for key, expected_value in expected_row.items():
            actual_value = actual_row.get(key)
            if actual_value != expected_value:
                mismatches.append(
                    {
                        "page": page,
                        "field": key,
                        "expected": expected_value,
                        "actual": actual_value,
                    }
                )

    if len(actual) != len(expected):
        mismatches.append({"field": "row_count", "expected": len(expected), "actual": len(actual)})

    print(json.dumps({"rows": len(actual), "mismatches": mismatches}, ensure_ascii=False, indent=2))
    return 1 if mismatches else 0


if __name__ == "__main__":
    raise SystemExit(main())
